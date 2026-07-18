from __future__ import annotations

import csv
from pathlib import Path

from src.platform.models.media_artifact import MediaArtifact
from src.platform.services.media_extractors.base import LocalExtractionInput


MAX_PREVIEW_ROWS = 10


class TableExtractor:
    name = "table_extractor"
    version = "1"
    supported_types = {"spreadsheet", "file"}

    def can_extract(self, artifact: MediaArtifact, path: Path) -> bool:
        suffix = path.suffix.lower()
        mime_type = (artifact.mime_type or "").lower()
        return suffix in {".csv", ".xlsx", ".xlsm"} or mime_type in {
            "text/csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }

    async def extract(self, payload: LocalExtractionInput):
        if payload.path.suffix.lower() == ".csv":
            return _extract_csv(payload)
        return _extract_xlsx(payload)


def _extract_csv(payload: LocalExtractionInput):
    from src.platform.services.media_ingestion import ExtractedNote

    rows: list[list[str]] = []
    with payload.path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for index, row in enumerate(reader):
            if index <= MAX_PREVIEW_ROWS:
                rows.append([str(cell)[:300] for cell in row])
            else:
                break
    headers = rows[0] if rows else []
    preview = rows[1:]
    text = _format_table_note(payload.path.name, "CSV", headers, preview, None)
    return ExtractedNote(
        title=f"表格笔记：{payload.artifact.original_name or payload.path.name}"[:180],
        summary=f"CSV 表格，列数 {len(headers)}，预览 {len(preview)} 行。",
        text=text,
        structured_data={"format": "csv", "headers": headers, "preview_rows": preview},
        source_url=payload.artifact.source_url or "",
        confidence=0.72 if headers else 0.35,
        warnings=[] if headers else ["CSV 没有可识别表头"],
    )


def _extract_xlsx(payload: LocalExtractionInput):
    from src.platform.services.media_ingestion import ExtractedNote

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl_not_installed") from exc

    workbook = load_workbook(payload.path, read_only=True, data_only=True)
    sheets = []
    text_parts = []
    for sheet in workbook.worksheets[:5]:
        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index > MAX_PREVIEW_ROWS:
                break
            rows.append(["" if cell is None else str(cell)[:300] for cell in row])
        headers = rows[0] if rows else []
        preview = rows[1:]
        sheet_data = {
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "headers": headers,
            "preview_rows": preview,
        }
        sheets.append(sheet_data)
        text_parts.append(_format_table_note(sheet.title, "XLSX", headers, preview, (sheet.max_row, sheet.max_column)))
    text = "\n\n".join(text_parts)
    return ExtractedNote(
        title=f"表格笔记：{payload.artifact.original_name or payload.path.name}"[:180],
        summary=f"Excel 表格，包含 {len(workbook.worksheets)} 个 sheet，已预览前 {len(sheets)} 个。",
        text=text[:20_000],
        structured_data={"format": "xlsx", "sheets": sheets},
        source_url=payload.artifact.source_url or "",
        confidence=0.74 if sheets else 0.35,
        warnings=[] if sheets else ["Excel 文件没有可读取工作表"],
    )


def _format_table_note(name: str, file_format: str, headers: list[str], preview: list[list[str]], shape) -> str:
    lines = [f"{file_format} 表格：{name}"]
    if shape:
        lines.append(f"规模：{shape[0]} 行 x {shape[1]} 列")
    if headers:
        lines.append("表头：" + " | ".join(headers))
    if preview:
        lines.append("预览：")
        lines.extend(" | ".join(row) for row in preview[:MAX_PREVIEW_ROWS])
    return "\n".join(lines)
